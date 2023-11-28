import openpyxl

def get_standard_standard_time_data():
    """
    读取C2到K4单元格中的数据,获取标准数据
    将数据存储到字典中
    """
    # 3种场景: 一般场景均值, 高耗场景均值, 二级界面均值
    scene = ['lowavgtime', 'highavgtime', 'towavgtime']
    # 标准数据
    standard_data = {}
    # 设备列表
    phone_list = []
    # 耗时数据（基准）
    standard_time_data = []

    # 获取设备列表
    for cell in sheet.iter_rows(min_row=1, max_row=1, min_col=3, max_col=11):
        phone_list = [cell.value for cell in cell]

    # 获取不同场景的UI耗时,将数据存储到standard_time_data
    for row in sheet.iter_rows(min_row=2, max_row=4, min_col=3, max_col=11):
        row_standard_time_data = [cell.value for cell in row]
        standard_time_data.append(row_standard_time_data)

    # 将不同数据与场景对应
    for i in range(3):
        dictionary = dict(zip(phone_list, standard_time_data[i]))
        standard_data[scene[i]] = dictionary


if __name__ == '__main__':
    # 打开Excel文件
    workbook = openpyxl.load_workbook('test.xlsx')

    # 选择工作表
    sheet = workbook['Sheet1']
    get_standard_standard_time_data()

    # 关闭Excel文件
    workbook.close()