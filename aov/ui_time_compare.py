import openpyxl
import xlwt
from time import sleep
# from thirdparty.wxwork_notice_api import *
import logging
import os

from openpyxl import load_workbook


# 读取 config/ui_time_standard.xlsx 中的标准耗时数据
def get_standard_time_consume(file_path):

    # 高耗界面均值
    V92_highavgtime = {}
    # 一般界面均值
    V92_lowavgtime = {}
    # 二级界面均值
    V92_towavgtime = {}

    # 打开工作簿
    workbook = load_workbook(filename=file_path)

    # 选择工作表
    sheet = workbook.active

    # 读取A1到K4的数据
    data = []
    for row in sheet.iter_rows(min_row=1, max_row=4, min_col=1, max_col=11):
        row_data = [cell.value for cell in row]
        data.append(row_data)

    # 将表格中的数据放入对应的字典中
    for _ in range(3):
        for i in range(2, 11):
            if data[1][0] == 'v92':
                if len(V92_highavgtime) != 9 and data[1][1] == '高耗界面均值':
                    V92_highavgtime[data[0][i]] = data[1][i]
                elif len(V92_lowavgtime) != 9 and data[2][1] == '一般界面均值':
                    V92_lowavgtime[data[0][i]] = data[2][i]
                elif len(V92_towavgtime) != 9 and data[3][1] == '二级界面均值':
                    V92_towavgtime[data[0][i]] = data[3][i]

    # 打印读取到的数据
    print(V92_highavgtime)
    print(V92_lowavgtime)
    print(V92_towavgtime)


V84_sumavgtime = {"oppo A33":"1.85","oppo A59":"1.49","vivo X9":"1.32","oppo r11":"1.1","xiaomi 8":"0.85",
           "iphone6":"2.29","iphone 6s":"1.1","iphone 7":"1.29","iphone 11 promax":"0.94"}

V91_sumavgtime = {"oppo A33":"2.26","oppo A59":"1.53","vivo X9":"1.47","oppo r11":"1.15","xiaomi 8":"0.75",
           "iphone6":"1.55","iphone 6s":"0.86","iphone 7":"0.79","iphone 11 promax":"0.96"}

V84_lowavgtime = {"oppo A33":"1.48","oppo A59":"1.23","vivo X9":"0.99","oppo r11":"0.89","xiaomi 8":"0.78",
           "iphone6":"1.55","iphone 6s":"0.97","iphone 7":"1.4","iphone 11 promax":"1.04"}

V91_lowavgtime = {"oppo A33":"1.53","oppo A59":"1.12","vivo X9":"1.1","oppo r11":"0.95","xiaomi 8":"0.66",
           "iphone6":"1.29","iphone 6s":"0.73","iphone 7":"0.64","iphone 11 promax":"0.99"}

V84_highavgtime = {"oppo A33":"3.55","oppo A59":"2.92","vivo X9":"2.37","oppo r11":"1.91","xiaomi 8":"1.51",
           "iphone6":"4.98","iphone 6s":"1.51","iphone 7":"2.37","iphone 11 promax":"1.36"}

V91_highavgtime = {"oppo A33":"4.45","oppo A59":"3.17","vivo X9":"2.95","oppo r11":"2.32","xiaomi 8":"1.39",
           "iphone6":"2.35","iphone 6s":"1.5","iphone 7":"1.33","iphone 11 promax":"1.6"}

V84_towavgtime = {"oppo A33":"1.54","oppo A59":"0.76","vivo X9":"0.89","oppo r11":"0.74","xiaomi 8":"0.51",
           "iphone6":"0.88","iphone 6s":"0.81","iphone 7":"0.77","iphone 11 promax":"0.56"}

V91_towavgtime = {"oppo A33":"0.82","oppo A59":"0.64","vivo X9":"0.84","oppo r11":"0.65","xiaomi 8":"0.48",
           "iphone6":"0.81","iphone 6s":"0.53","iphone 7":"0.48","iphone 11 promax":"0.64"}

logging.basicConfig(level=logging.DEBUG,
                            # 设置输出格式，年月日分秒毫秒，行数，等级名，打印的信息
                            format="%(asctime)s - %(filename)s[line:%(lineno)d] - %(levelname)s: %(message)s")

#高耗界面UI耗时对比
def highavgtime_msg (webhook_key,path):
    # wx_tool = WxWorkNotice(webhook_key)
    #打开需要对比数据的路径
    contrast_excel = openpyxl.load_workbook(path)
    contrast_sheet = contrast_excel["Sheet1"]
    for i in contrast_sheet:
        for j in i:
            print(type(j.value))
    for phone_name,phone_time in V91_highavgtime.items() :
        logging.info(f"phonename:{phone_name},phonetime:{phone_time}")
    # report_txt = "测试测试"
    # wx_tool.send_txt_mgs(report_txt)
    # logging.info("已通知在相关群")

if __name__ == '__main__':
    # webhook_key = "d6c81a89-8924-4caf-b9f8-d354311f3df2"
    # path = "F:\\国内性能数据\\test.xlsx"
    # highavgtime_msg(webhook_key, path)
    get_standard_time_consume('config/ui_time_standard.xlsx')