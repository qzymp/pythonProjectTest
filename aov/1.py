import openpyxl
import xlwt
from time import sleep
import logging
import os
'''
使用说明
1.将perfdog excel数据以 xxx_perfdog网址编号  的格式命名
2.将存放perfdog excel数据的路径放置 main中直接运行即可
ps:数据获取时切记不要打开对应的perfdog excel数据，否则会报错，
且产出的目录默认设定咸了在D盘的根目录下，如需要更改，可直接修改 savepath 参数即可
'''
# 常规数据
# 常规数据--all--标签下需要遍历的字段
perfdogfield_alllist = ["Avg(Memory)[MB]","Avg(Memory+Swap)[MB]","Peak(Memory)[MB]","Peak(Memory+Swap)[MB]"]
perfdogfield_allendlist = ["VirtualMemory[MB]"]

# 常规数据--游戏中--标签下需要遍历的字段
perfdogfield_gamelist = ["Avg(FPS)","FPS>=18[%]","Var(FPS)","Min(FPS)","Jank(/10min)","BigJank(/10min)",
                        "Avg(Memory)[MB]","Avg(Memory+Swap)[MB]","Peak(Memory)[MB]","Peak(Memory+Swap)[MB]",
                         "Avg(AppCPU)[%]","(Recv+Send)[KB/10min]","Avg(Power)[mW]","FPower[mW]","Sum(Battery)[mWh]"]
perfdogfield_gameendlist = ["TotalCPU[%]","GUsage[%]","Voltage[mV]","Current[mA]","Recv[KB/s]","Send[KB/s]"]

# 常规数据除了--all--标签外，需要遍历的标签
# gamelabel_list = ['首次进排行榜拍脸','设置说明窗口','打开排行榜首页','巅峰榜战区','巅峰榜全国','巅峰榜右上角说明','巅峰榜右下角说明','切换榜单','天梯榜-全国','皮肤榜-全国','贵族积分榜-全国','战令榜-全国','成就榜-全国','胜场榜-全国','连续胜场榜-全国','人气榜-全国','名师榜-全国']
# gamelabel_list = ['大厅','大厅邀请','房间','房间小浮窗','房间邀请','结算','结算小浮窗','结算邀请','结算邀请按钮']
# gamelabel_list = ['英雄榜','分路榜','亲密关系榜']
# gamelabel_list = ['皮肤展示','all']
# gamelabel_list = ['all']
# gamelabel_list = ['大厅-新版']
# gamelabel_list = ['第一局','第二局','第三局']
# gamelabel_list = ['第一关','第二关','第三关','第四关','第五关']
# gamelabel_list = ['组队房间-无灵花','组队房间-灵花','组队房间-灵宝','灵宝大厅','大厅无灵花无灵宝']
# gamelabel_list = ['1','2','3','4','5','6','7','8','9','10','11','12','13','14','15','16',]
# gamelabel_list = ['第1波','第2波','第3波','第4波','第5波','第6波','第7波','第8波','第9波','第10波','第11波','第12波','第13波','第14波','第15波','第16波']

# gamelabel_list = ['简单怪物120','简单怪物130','简单怪物200',
#                 '中等怪物45','中等怪物50','中等怪物33','中等怪物34',
# '复杂怪物13','复杂怪物14','复杂怪物110','复杂怪物120',
# '简单特效45','简单特效50','简单特效300',
# '中等特效4','中等特效5','中等特效60',
# '复杂特效2','复杂特效3','复杂特效19','复杂特效20','空场景']


# gamelabel_list = ['简单怪物500','简单怪物550','简单怪物1500',
#                 '中等怪物300','中等怪物310','中等怪物260','中等怪物270',
# '复杂怪物200','复杂怪物210','复杂怪物210','复杂怪物220',
# '简单特效650','简单特效700','简单特效1400','简单特效1500',
# '中等特效40','中等特效45','中等特效290','中等特效300',
# '复杂特效16','复杂特效17','复杂特效70','复杂特效80','空场景']
# gamelabel_list = ['战令界面','战令升级','兑换界面','甄宝界面']
# gamelabel_list = ['战令-道具','兑换','甄宝','战令-皮肤','','']
# gamelabel_list = ['进入排位','想玩英雄设置窗口（熟练度排序）','想玩设置窗口（熟练度）','想玩英雄设置窗口 (练度排序.','想玩英雄设置窗口（战力排序）','想玩英雄设置窗口（战力排序）','巅峰赛','巅峰赛想玩英雄','巅峰想玩英雄','巅峰赛想玩的英雄','荣耀记录']
# gamelabel_list = ['0','30','2','4','10','20']
# gamelabel_list = ['局内']
# gamelabel_list = ['想玩英雄设置窗口（战力排序）']
# gamelabel_list = ['局内']
# gamelabel_list = ['局内','第二局','第三局']
# gamelabel_list = ['弹出玩法引导窗口','接收本赛季礼包界面','获得选手卡片界面','点击‘去养成’','点击‘去进阶’','选中金卡-强化','进阶失败','进阶成功','进阶失败点击知道了','点击上方收集材料的‘+’','点击强化','升星成功','升星失败','选中金卡-5星选手卡','积分榜','积分榜上下滑动30s','点击整条排行榜','点击未开榜','','','','','','','','','','','','','','']
# gamelabel_list = ['','','','','','','','']
# gamelabel_list = ['梦幻组合界面','赛季礼包界面','选手卡片界面','去养成','去进阶','选中金卡-强化','进阶失败','进阶成功','任务没做进阶失败','收集材料的‘+’','货币足够，选中第一个强化项','升星成功','升星失败','金卡-5星选手卡','记录首次进入积分','滑动30s','大卡片','未开榜','','','','','','','','','','','','','','']
# gamelabel_list = ['排位赛大厅','想玩英雄-熟练度','想玩英雄-战力','荣耀记录','巅峰赛','巅峰赛-想玩英雄']
# gamelabel_list = ['基地']
# gamelabel_list = ['粉色聊天界面','蓝色聊天框']
gamelabel_list = ['5经验']
# gamelabel_list = ['孙尚香','李白','孙策']
# gamelabel_list = ['大厅1','KPL标清全屏','KPL高清全屏','KPL超清全屏','KPL蓝光全屏','KPL蓝光半屏','KPL小兵对决','K甲蓝光全屏','K甲蓝光半屏','基地','我的','视频','观战','大厅2','战令','竞猜','活动','带货','曲线','红蓝弹幕','天使之翼','AG捧杯']
# gamelabel_list = ['孙策','李白','孙尚香']
# gamelabel_list = ['引导','礼包','获得卡片']
# gamelabel_list = ['1','2','3','4']
# gamelabel_list = ['活动主页','兑换商店主页','兑换商店首个物品','奖池主页','奖池第二个物品','令牌赠送','单抽1','单抽2','十连1','十连2','抽奖记录']
# gamelabel_list = ['局1|游戏中','局2|游戏中','局3|游戏中']
# gamelabel_list = ['单抽1','单抽2','十连1','十连2']
# gamelabel_list = ['凯撒团','魔龙团','暗影凯撒团','凯撒周边团','魔龙团2','魔神凯撒团']
# gamelabel_list = ['局内','AOV合入英雄第一局-绮罗','AOV合入英雄第二局-皮皮','AOV合入英雄第三局-曹操']
# gamelabel_list = ['复杂特效13','复杂特效14']
# gamelabel_list = ['全民电竞','全国大赛邀请','电竞商店','电竞福利','我的生涯','排行榜','赛事选择','主题赛','电竞徽章','近期赛事']


# ui标签
perfdoguilabel_list = ["个人资料1","个人资料2","战令1","战令2","聊天1","聊天2","测试测试测试",
                       "英雄1","英雄2","定制化1","定制化2","备战1","备战2",
                       "背包1","背包2","好友1","好友2","下载1","下载2",
                       "设置1","设置2","邮件1","邮件2","排行榜1","排行榜2",
                       "充值1","充值2","首充1","首充2","活动1","活动2"]

# ui数据--每个标签--下字段需要遍历的字段
perfdoguifield_list = ["Avg(FPS)","FPS>=18[%]","Jank(/10min)","Peak(Memory)[MB]","Avg(AppCPU)[%]","Peak(Memory+Swap)[MB]"]
perfdoguifield_endlist = ["SwapMemory[MB]","GUsage[%]"]

# 常规数据获取
def catch_routineperfdog_data(path):
    try:
        logging.basicConfig(level=logging.DEBUG,
                            # 设置输出格式，年月日分秒毫秒，行数，等级名，打印的信息
                            format="%(asctime)s - %(filename)s[line:%(lineno)d] - %(levelname)s: %(message)s")
        perdog_files = os.listdir(path)
        perfdog_list = []
        for perfdog_file in perdog_files:
            perfdogfile_name = perfdog_file.split(".")
            perfdogfile_id = perfdogfile_name[0].split("_")
            perfdog_url = "https://perfdog.qq.com/case_detail/"+perfdogfile_id[1]
            http_perfdog = '=HYPERLINK("{}","{}")'.format(perfdog_url,'链接')
            perfdog_datas = openpyxl.load_workbook(filename=path+"\\"+perfdog_file)
            logging.info("正在获取"+perfdog_file+"的全局的数据...")

            # 获取 all 标签的数据   all 标签下的数据对比于其它的标签的取法有不同之处，因此逻辑单独处理
            perfdog_data_all = perfdog_datas['all']
            logging.info("----------------------------------------------------------")
            perfdog_data_all_maxrow = perfdog_data_all.max_row

            # 获取 all 标签里对应数据的下标
            all_row_range = perfdog_data_all[8:9]

            perfdog_all_data_list = []
            perfdog_all_updata = {}
            for all_row in all_row_range:
                for all_cell in all_row:
                    perfdog_all_data_list.append(all_cell.value)
                for perfall_field in perfdogfield_alllist:
                    if perfall_field in perfdog_all_data_list:
                        allfirst_index = perfdog_all_data_list.index(perfall_field)
                        allfirst_column = allfirst_index + 1
                        allfirst_value = perfdog_data_all.cell(9,allfirst_column).value
                        perfdog_all_updata.setdefault(perfall_field,allfirst_value)
                    else:
                        logging.error(perfall_field+"字段不存在")
                        perfdog_all_updata.setdefault(perfall_field,"/")

            all_row_end_range = perfdog_data_all[perfdog_data_all_maxrow - 2:perfdog_data_all_maxrow]
            perfdog_data_all_end_list = []
            perfdog_all_enddata = {}
            for all_end_row in all_row_end_range:
                for all_end_cell in all_end_row:
                    perfdog_data_all_end_list.append(all_end_cell.value)
                for perfall_endfield in perfdogfield_allendlist:
                    if perfall_endfield in perfdog_data_all_end_list:
                        allend_index = perfdog_data_all_end_list.index(perfall_endfield)
                        allend_column = allend_index + 1
                        if perfall_endfield == "VirtualMemory[MB]":   # 需要获取最大值的参数
                            allend_value = perfdog_data_all.cell(perfdog_data_all_maxrow,allend_column).value
                        else:                                       # 平均值获取
                            allend_value = perfdog_data_all.cell(perfdog_data_all_maxrow-1,allend_column).value
                        perfdog_all_enddata.setdefault(perfall_endfield,allend_value)
                    else:
                        logging.error(perfall_endfield+"字段不存在")
                        perfdog_all_enddata.setdefault(perfall_endfield,"/")

            # 将获取到的字段值赋值
            all_avg_memory = perfdog_all_updata["Avg(Memory)[MB]"]
            all_avg_memory_swap = perfdog_all_updata["Avg(Memory+Swap)[MB]"]
            all_peak_memory = perfdog_all_updata["Peak(Memory)[MB]"]
            all_peak_memory_swap = perfdog_all_updata["Peak(Memory+Swap)[MB]"]
            all_peak_virtual_memory = perfdog_all_enddata["VirtualMemory[MB]"]

            logging.info('全局数据获取完成')

            logging.info("正在获取"+perfdog_file+"game1的数据...")

            # 获取 局内 标签的数据

            perfdog_label_list = []
            for perfdog_label_data in perfdog_datas.worksheets:
                perfdog_label_list.append(perfdog_label_data.title)
            for perfdoggamelabel in gamelabel_list:
                if perfdoggamelabel not in perfdog_label_list:
                    logging.error(perfdogfile_name[0] + "文件中没有" + perfdoggamelabel + "这个sheet")
                else:
                    perfdog_data_game1 = perfdog_datas[perfdoggamelabel]  # 填入自己数据里对局1的标签名
                    # 获取 局内 标签的最大行
                    perfdog_data_game1_maxrow = perfdog_data_game1.max_row

                    # 局内 标签里对应数据的下标
                    game1_row_range = perfdog_data_game1[8:9]
                    perfdog_data_game1_list = []
                    perfdog_game_updata = {}
                    for game1_row in game1_row_range:
                        for game1_cell in game1_row:
                            perfdog_data_game1_list.append(game1_cell.value)
                        for perfgame_field in perfdogfield_gamelist:
                            if perfgame_field in perfdog_data_game1_list:
                                game_index = perfdog_data_game1_list.index(perfgame_field)
                                game_column = game_index + 1
                                game_value = perfdog_data_game1.cell(9,game_column).value
                                perfdog_game_updata.setdefault(perfgame_field,game_value)
                            else:
                                logging.error(perfgame_field+"字段不存在")
                                perfdog_game_updata.setdefault(perfgame_field,"/")

                    game1_row_end_range = perfdog_data_game1[perfdog_data_game1_maxrow - 2:perfdog_data_game1_maxrow]
                    perfdog_data_game1_end_list = []
                    perfdog_game_enddata = {}
                    for game1_row_end in game1_row_end_range:
                        for game1_cell_end in game1_row_end:
                            perfdog_data_game1_end_list.append(game1_cell_end.value)
                        for perfgame_endfield in perfdogfield_gameendlist:
                            if perfgame_endfield in perfdog_data_game1_end_list:
                                gameend_index = perfdog_data_game1_end_list.index(perfgame_endfield)
                                gameend_column = gameend_index + 1
                                gameend_value = perfdog_data_game1.cell(perfdog_data_game1_maxrow - 1,gameend_column).value
                                perfdog_game_enddata.setdefault(perfgame_endfield,gameend_value)
                            else:
                                logging.error(perfgame_endfield+"字段不存在")
                                perfdog_game_enddata.setdefault(perfgame_endfield,"/")

                    # 将获取到的字段值赋值
                    game1_avg_fps = perfdog_game_updata["Avg(FPS)"]
                    game1_avg_f18 = perfdog_game_updata["FPS>=18[%]"]
                    game1_varfos = perfdog_game_updata["Var(FPS)"]
                    game1_minfps = perfdog_game_updata["Min(FPS)"]
                    game1_jank = perfdog_game_updata["Jank(/10min)"]
                    game1_bigjank = perfdog_game_updata["BigJank(/10min)"]
                    game1_avg_memory = perfdog_game_updata["Avg(Memory)[MB]"]
                    game1_avg_memory_swap = perfdog_game_updata["Avg(Memory+Swap)[MB]"]
                    game1_peak_memory = perfdog_game_updata["Peak(Memory)[MB]"]
                    game1_peak_memory_swap = perfdog_game_updata["Peak(Memory+Swap)[MB]"]
                    game1_avg_cpu = perfdog_game_updata["Avg(AppCPU)[%]"]
                    game1_avg_total_cpu = perfdog_game_enddata["TotalCPU[%]"]
                    game1_avggpu_data = perfdog_game_enddata["GUsage[%]"]
                    if game1_avggpu_data == "/" or game1_avggpu_data < 0:
                        game1_gpu_data = "/"
                    else:
                        game1_avggpu_percentage_data = game1_avggpu_data / 100
                        game1_gpu_data = format(game1_avggpu_percentage_data, '.1%')

                    game1_recvkbs = perfdog_game_enddata["Recv[KB/s]"]
                    if game1_recvkbs != "/":
                        game1_recvkbm = game1_recvkbs * 600
                    else:
                        game1_recvkbm = "/"
                    game1_sendkbs = perfdog_game_enddata["Send[KB/s]"]
                    if game1_sendkbs != "/":
                        game1_sendkbm = game1_sendkbs * 600
                    else:
                        game1_sendkbm = "/"
                    game1_avg_power = perfdog_game_updata["Avg(Power)[mW]"]
                    game1_f_power = perfdog_game_updata["FPower[mW]"]
                    game1_sum_battery = perfdog_game_updata["Sum(Battery)[mWh]"]
                    game1_voltage = perfdog_game_enddata["Voltage[mV]"]
                    game1_current = perfdog_game_enddata["Current[mA]"]

                    logging.info('game1数据获取完成...')

                    # 将部分数据处理成百分数
                    if game1_avg_f18 != "/":
                        game1_avg_f18_data = game1_avg_f18 / 100
                        game1_avg_f18_data_save = format(game1_avg_f18_data, '.1%')
                    else:
                        game1_avg_f18_data_save = "/"

                    if game1_avg_cpu != "/":
                        game1_avg_cpu_data = game1_avg_cpu / 100
                        game1_avg_cpu_data_save = format(game1_avg_cpu_data, '.1%')
                    else:
                        game1_avg_cpu_data_save = "/"

                    if game1_avg_total_cpu != "/":
                        game1_avg_total_cpu_data = game1_avg_total_cpu / 100
                        game1_avg_total_cpu_data_save = format(game1_avg_total_cpu_data, '.1%')
                    else:
                            game1_avg_total_cpu_data_save = "/"

                    # 将抓取的数据传入list
                    perfdog_list.append([perfdogfile_id[0],perfdoggamelabel,game1_avg_fps,game1_avg_f18_data_save,game1_varfos,game1_minfps,game1_jank,game1_bigjank,all_avg_memory,
                                 all_avg_memory_swap,all_peak_memory,all_peak_memory_swap,all_peak_virtual_memory,
                                 game1_avg_memory,game1_avg_memory_swap,game1_peak_memory,game1_peak_memory_swap,
                                 game1_avg_cpu_data_save,game1_avg_total_cpu_data_save,game1_gpu_data,game1_recvkbm,game1_sendkbm,
                                 game1_avg_power,game1_f_power,game1_sum_battery,game1_voltage,game1_current,http_perfdog])

    except Exception as e:
        logging.error(e)
    finally:
        perfdog_datas.close()
        logging.info("成功关闭打开的文件")

    # 产出excel表格
    logging.info('正在输出至excel中...')
    # 创建excel表格类型文件
    workbook = xlwt.Workbook(encoding='utf-8', style_compression=0)
    # 在excel表格类型文件中建立一张sheet表单
    sheet = workbook.add_sheet('test', cell_overwrite_ok=True)
    # 自定义列名
    col = ('获取文件','文件标签','avgFPS', '>18帧占比','Var(FPS)', 'Min(FPS)', 'Jank', 'BigJank', '整局PSS均值', '整局Avg(PSS+Swap)', '整局MaxPSS', '整局Max(PSS+Swap)',
            '整局虚拟内存峰值', 'PSS均值', 'Avg(PSS+Swap)', 'MaxPSS', 'Max(PSS+Swap)', 'AvgCPU', 'AvgToTalCPU','GPU',
           'recv/10min','send/10min','Avg(Power)','F(Power)','Sum(Battery)','Voltage','Current','perfdog链接')

    # 将列属性元组col写进sheet表单中
    for i in range(0,len(col)):
        sheet.write(0, i, col[i])
    # 将数据写进sheet表单中
    datalist = []
    for datalists in perfdog_list:
        datalist.append(datalists)
    for i in range(0, len(datalist)):
        data = datalist[i]
        for j in range(0, len(col)):
            sheet.write(i + 1, j, data[j])
    # 保存excel文件
    savepath = path + '\\test.xls'
    workbook.save(savepath)

    logging.info("输出完成")

# ui数据获取
def catch_uiperfdog_data(path):
    try:
        logging.basicConfig(level=logging.DEBUG,
                            # 设置输出格式，年月日分秒毫秒，行数，等级名，打印的信息
                            format="%(asctime)s - %(filename)s[line:%(lineno)d] - %(levelname)s: %(message)s")
        perdog_files = os.listdir(path)
        perfdog_list = []
        for perfdog_file in perdog_files:
            perfdogfile_name = perfdog_file.split(".")
            perfdogfile_id = perfdogfile_name[0].split("_")
            perfdog_url = "https://perfdog.qq.com/case_detail/"+perfdogfile_id[1]
            perfdog_datas = openpyxl.load_workbook(filename=path+"\\"+perfdog_file)
            logging.info("正在获取"+perfdog_file+"的全局的数据...")

            # 获取 UI 标签的数据
            perfdog_label_list = []
            for perfdog_label_data in  perfdog_datas.worksheets:
                perfdog_label_list.append(perfdog_label_data.title)
            for perfdoguilabel in perfdoguilabel_list:
                if perfdoguilabel not in perfdog_label_list:
                    logging.error(perfdogfile_name[0]+"文件中没有"+perfdoguilabel+"这个sheet")
                else:
                    perfdog_data_all = perfdog_datas[perfdoguilabel]
                    perfdog_data_all_maxrow = perfdog_data_all.max_row

                    # 获取各个标签里对应数据的下标
                    all_row_range = perfdog_data_all[8:9]
                    perfdog_all_data_list = []
                    perfdog_ui_updata = {}
                    for all_row in all_row_range:
                        for all_cell in all_row:
                            perfdog_all_data_list.append(all_cell.value)
                        for perfdoguifield in perfdoguifield_list:
                            if perfdoguifield in perfdog_all_data_list:
                                allfirst_index = perfdog_all_data_list.index(perfdoguifield)
                                allfirst_column = allfirst_index + 1
                                allfirst_value = perfdog_data_all.cell(9,allfirst_column).value
                                perfdog_ui_updata.setdefault(perfdoguifield,allfirst_value)
                            else:
                                logging.error(perfdoguifield+"字段不存在")
                                perfdog_ui_updata.setdefault(perfdoguifield,"/")

                    all_row_end_range = perfdog_data_all[perfdog_data_all_maxrow - 2:perfdog_data_all_maxrow]
                    perfdog_data_all_end_list = []
                    perfdog_ui_enddata = {}
                    for all_end_row in all_row_end_range:
                        for all_end_cell in all_end_row:
                            perfdog_data_all_end_list.append(all_end_cell.value)
                        for perfdoguiendfield in perfdoguifield_endlist:
                            if perfdoguiendfield in perfdog_data_all_end_list:
                                allfirst_index = perfdog_data_all_end_list.index(perfdoguiendfield)
                                allfirst_column = allfirst_index + 1
                                allfirst_value = perfdog_data_all.cell(perfdog_data_all_maxrow, allfirst_column).value
                                perfdog_ui_enddata.setdefault(perfdoguiendfield, allfirst_value)
                            else:
                                logging.error(perfdoguiendfield + "字段不存在")
                                perfdog_ui_enddata.setdefault(perfdoguiendfield, "/")

                    # 将获取到的字段值赋值
                    ui_avg_fps = perfdog_ui_updata["Avg(FPS)"]
                    ui_avg_f18 = perfdog_ui_updata["FPS>=18[%]"]
                    ui_jank = perfdog_ui_updata["Jank(/10min)"]
                    ui_peak_memory = perfdog_ui_updata["Peak(Memory)[MB]"]
                    ui_memory_swap = perfdog_ui_enddata["SwapMemory[MB]"]
                    if ui_memory_swap == "/":
                        ui_peak_memory_swap = ui_peak_memory
                    else:
                        ui_peak_memory_swap = perfdog_ui_updata["Peak(Memory+Swap)[MB]"]
                    ui_avg_cpu = perfdog_ui_updata["Avg(AppCPU)[%]"]
                    ui_avggpu_data = perfdog_ui_enddata["GUsage[%]"]

                    logging.info(perfdoguilabel+"数据获取完成")

                    # 将部分数据处理成百分数
                    if ui_avg_f18 != "/":
                        ui_avg_f18_data = ui_avg_f18 / 100
                        ui_avg_f18_data_save = format(ui_avg_f18_data, '.1%')
                    else:
                        ui_avg_f18_data_save = "/"
                    if ui_avg_cpu != "/":
                        ui_avg_cpu_data = ui_avg_cpu / 100
                        ui_avg_cpu_data_save = format(ui_avg_cpu_data, '.1%')
                    else:
                        ui_avg_cpu_data_save = "/"
                    if ui_avggpu_data != "/":
                        ui_avg_gpu_data = ui_avggpu_data / 100
                        ui_avg_gpu_data_save = format(ui_avg_gpu_data, '.1%')
                    else:
                        ui_avg_gpu_data_save = "/"

                    # 将抓取的数据传入list
                    perfdog_list.append([perfdogfile_name[0],perfdoguilabel,ui_avg_fps,ui_avg_f18_data_save,ui_jank,
                                     ui_peak_memory_swap,ui_avg_cpu_data_save,ui_avg_gpu_data_save,perfdog_url])

    except Exception as e:
        logging.error(e)
    finally:
        perfdog_datas.close()
        logging.info("成功关闭打开的文件")

    # 产出excel表格
    logging.info('正在输出至excel中...')
    # 创建excel表格类型文件
    workbook = xlwt.Workbook(encoding='utf-8', style_compression=0)
    # 在excel表格类型文件中建立一张sheet表单
    sheet = workbook.add_sheet('test', cell_overwrite_ok=True)
    # 自定义列名
    col = ('设备名','标签名','avgFPS', '>18帧占比','Jank','PeakMemory', 'AvgCPU', 'GPU','perfdog链接')

    # 将列属性元组col写进sheet表单中
    for i in range(0,len(col)):
        sheet.write(0, i, col[i])
    # 将数据写进sheet表单中
    datalist = []
    for datalists in perfdog_list:
        datalist.append(datalists)
    for i in range(0, len(datalist)):
        data = datalist[i]
        for j in range(0, len(col)):
            sheet.write(i + 1, j, data[j])
    # 保存excel文件
    savepath = 'D:\\test.xls'
    workbook.save(savepath)

    logging.info("输出完成")

if __name__ == '__main__':
    # 常规数据 + 功耗数据 获取
    catch_routineperfdog_data("D:\\perfdog_data")
    # ui数据获取
    # catch_uiperfdog_data("D:\\海外性能测试\\uitest")
